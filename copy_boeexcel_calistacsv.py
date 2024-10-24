r"""
Copy results from BoE excel to Calista csv.

!!! Although the code is tested, it is adviced to manually cross check the transferred results. !!!

Author: Ahsan Habib, School of IT, Deakin.
"""
# !pip install openpyxl

def install_module(module_name):
    import sys
    import subprocess
    # implement pip as a subprocess:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', module_name])

install_module('openpyxl')
install_module('pandas')

import openpyxl
import pandas as pd


def find_col_by_name(sheet, header_row_num, header_name):
    col = 1
    while True:
        cell_obj = sheet.cell(row=header_row_num, column=col)
        if cell_obj.value == header_name:
            return col
        col += 1


def main():
    "Load xlsx file."
    src_wb = openpyxl.load_workbook(SRC_XLSX_FILE, data_only=True)
    try:
        src_sheet_obj = src_wb[SRC_WORKBOOK]
    except:
        raise Exception(f"Source ({SRC_XLSX_FILE}) does not have workbook '{SRC_WORKBOOK}'")
    
    src_row = SRC_DATA_START

    "Read columns - ID, Final and Result"
    src_col_studentid = find_col_by_name(src_sheet_obj, SRC_DATA_START-1, "ID")
    if src_col_studentid is None:
        raise Exception(f"Expected column 'ID' not found.")

    src_col_final = find_col_by_name(src_sheet_obj, SRC_DATA_START-1, "Final")
    if src_col_final is None:
        raise Exception(f"Expected column 'Final' not found.")

    src_col_result = find_col_by_name(src_sheet_obj, SRC_DATA_START-1, "Result")
    if src_col_result is None:
        raise Exception(f"Expected column 'Result' not found.")

    "Read Final and Grade columns and store in buffer."
    studentid_total_map = {}
    while True:
        "Read ID"
        cell_obj = src_sheet_obj.cell(row=src_row, column=src_col_studentid)
        if cell_obj is None or cell_obj.value is None:
            break
        src_id = int(cell_obj.value)

        "Read Final, allow None value"
        cell_obj = src_sheet_obj.cell(row=src_row, column=src_col_final)
        col_final = cell_obj
        # if cell_obj is None or cell_obj.value is None:
        #     break

        "Read Result, allow None value"
        cell_obj = src_sheet_obj.cell(row=src_row, column=src_col_result)
        col_result = cell_obj

        studentid_total_map[int(src_id)] = [col_final.value, col_result.value]
        src_row += 1

    "Now, read destination Calista CSV file and fill data."
    df = pd.read_csv(DST_CSV_FILE, encoding='latin-1')
    
    ids_not_found = []
    for index, row in df.iterrows():
        std_id = int(row['Person ID'])
        # print(std_id)

        if studentid_total_map.get(std_id) is None:
            ids_not_found.append(std_id)
            continue
        df.loc[index, 'Mark'] = studentid_total_map.get(std_id)[0]
        df.loc[index, 'Grade'] = studentid_total_map.get(std_id)[1]
    print(f"{len(ids_not_found)} IDs not found; => {ids_not_found}")

    "Save to CSV file"
    output_file = f"{DST_CSV_FILE.split('.')[0]}_updated.csv"
    df.to_csv(output_file, encoding='latin-1', index=False)
    print(f"\nSaved to: {output_file}")


if __name__ == "__main__":
    SRC_DATA_START = 13
    SRC_XLSX_FILE = "results/SIT225-2024-T2.xlsx"
    SRC_WORKBOOK = "Results"

    "Source parameters"
    SRC_XLSX_FILE = input("Source xlsx file: ")
    SRC_WORKBOOK = input("Source workbook: ")
    
    DST_CSV_FILE = input("Destination Calista CSV file: ")
    
    print("\nProcessing ...\n")
    
    main()