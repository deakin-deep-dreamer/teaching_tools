r"""
Copy results from OnTrack export xlsx to BoE result xlsx.

!!! Although the code is tested, it is adviced to manually cross check the transferred results. !!!

Author: Ahsan Habib, School of IT, Deakin.
"""
# !pip install openpyxl

def install_openpyxl():
    import sys
    import subprocess
    # implement pip as a subprocess:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
install_openpyxl()

import openpyxl

def lookup_src(student_id, src_sheet, id_col=3, val_col=4, data_start_row=2):
    row_num = data_start_row
    while True:
        cell_obj = src_sheet.cell(row=row_num, column=id_col)
        # print(cell_obj, cell_obj.value)
        if cell_obj is None or cell_obj.value is None:
            # print("lookup_src cell value is None")
            return
        if int(cell_obj.value) == student_id:
            # print(f"Found ID:{student_id} at row {row_num}")
            # get the grade
            cell_obj = src_sheet.cell(row=row_num, column=val_col)            
            # return round(float(cell_obj.value), 0)
            return cell_obj.value
        row_num += 1


def find_col_by_name(sheet, header_row_num, header_name):
    col = 1
    while True:
        cell_obj = sheet.cell(row=header_row_num, column=col)
        if cell_obj.value == header_name:
            return col
        col += 1


def main():
    dest_wb = openpyxl.load_workbook(DEST_XLSX_FILE)
    try:
        dest_sheet_obj = dest_wb[DEST_WORKBOOK]
    except:
        raise Exception(f"Destination ({DEST_XLSX_FILE}) does not have workbook '{DEST_WORKBOOK}'")

    src_wb = openpyxl.load_workbook(SRC_XLSX_FILE)
    try:
        src_sheet_obj = src_wb[SRC_WORKBOOK]
    except:
        raise Exception(f"Source ({SRC_XLSX_FILE}) does not have workbook '{SRC_WORKBOOK}'")

    dest_col_notes = find_col_by_name(dest_sheet_obj, DEST_DATA_START-1, "Notes")
    if dest_col_notes is None:
        print(f"Notes column not found in destination header row {DEST_DATA_START-1}")
    else:
        print("Notes:", dest_col_notes)

    dst_update_count = 0
    dst_row = DEST_DATA_START
    dst_missing_ids = []
    while True:
        cell_obj = dest_sheet_obj.cell(row=dst_row, column=DEST_ID_COL)
        if cell_obj is None or cell_obj.value is None:
            break
        dst_id = int(cell_obj.value)
        # print(dst_id)    

        found = lookup_src(dst_id, src_sheet_obj, id_col=SRC_ID_COL, val_col=SRC_MARK_COL)
        if found is not None:
            dest_sheet_obj.cell(row=dst_row, column=DEST_MARK_COL).value = round(float(found), 0)
            dst_update_count += 1

            # Find comments
            found = lookup_src(dst_id, src_sheet_obj, id_col=SRC_ID_COL, val_col=SRC_COMMENT_COL)
            if found is not None:
                dest_sheet_obj.cell(row=dst_row, column=dest_col_notes).value = found
        else:
            dst_missing_ids.append(dst_id)
        dst_row += 1

    dst_total_count = dst_row - DEST_DATA_START
    print(f"Total {dst_update_count} rows updated out of {dst_total_count}, missing:{dst_total_count-dst_update_count}.")
    print(f"Missing IDs: {dst_missing_ids}")
    output_file = f"{DEST_XLSX_FILE.split('.')[0]}-updated.xlsx"
    dest_wb.save(output_file)
    print(f"Output saved to {output_file}")


if __name__ == "__main__":
    DEST_DATA_START = 13
    DEST_ID_COL = 3
    DEST_MARK_COL = 7
    DEST_XLSX_FILE = "results/SIT225-2024-T2.xlsx"
    DEST_WORKBOOK = "Results"

    SRC_XLSX_FILE = "results/SIT225-Students-ontrack-all.xlsx"
    SRC_WORKBOOK = "SIT225-Students-ontrack-all"
    SRC_ID_COL = 3
    SRC_MARK_COL = 4
    SRC_COMMENT_COL = 5

    "Destination parameters"
    DEST_XLSX_FILE = input("Destination xlsx file: ")
    DEST_WORKBOOK = input("Destination workbook: ")
    try:
        DEST_ID_COL = int(input("Destination student-ID column-number (default 3): "))
    except ValueError:
        DEST_ID_COL = 3
    try:
        DEST_MARK_COL = int(input("Destination grade/mark column-number (default 7): "))
    except ValueError:
        DEST_MARK_COL = 7
    try:
        DEST_DATA_START = int(input("Destination data-row start (default 13): "))
    except ValueError:
        DEST_DATA_START = 13
    print()

    SRC_XLSX_FILE = input("Source xlsx file: ")
    SRC_WORKBOOK = input("Source workbook: ")
    try:
        SRC_ID_COL = int(input("Source student-ID column-number (default 3): "))
    except ValueError:
        SRC_ID_COL = 3
    try:
        SRC_MARK_COL = int(input("Source grade/mark column-number (default 4): "))
    except ValueError:
        SRC_MARK_COL = 4
    try:
        SRC_COMMENT_COL = int(input("Source comment column-number (default 5): "))
    except ValueError:
        SRC_COMMENT_COL = 5
    
    print("\nProcessing...\n")
    main()
    